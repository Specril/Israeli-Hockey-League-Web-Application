import pandas as pd
from openpyxl import load_workbook
import json
import numpy as np
import datetime

# Define a custom JSON encoder
class NpEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (np.integer, np.int64)):
            return int(obj)
        elif isinstance(obj, (np.floating, np.float64)):
            return int(obj) if obj.is_integer() else float(obj)
        elif isinstance(obj, (pd.Timestamp, datetime.datetime)):
            return obj.isoformat()
        elif isinstance(obj, (pd.Timedelta, pd.Period)):
            return str(obj)
        return super().default(obj)

# Function to create second table from first table
def create_second_table(first_table):
    second_table_rows = []
    for _, row in first_table.iterrows():
        player_id = row['Player_ID']
        team_id = row['Team_ID']
        for i in range(1, (len(row) - 2) // 2 + 1):
            game_id = row[f'Game_ID {i}']
            num_goals = int(row[f'Game_Goals {i}'])
            for goal_id in range(num_goals):
                second_table_rows.append({
                    'Goal_ID': f'{player_id}_{game_id}_{goal_id + 1}',
                    'Game_ID': game_id,
                    'Player_ID': player_id,
                    'Team_ID': team_id,
                    'type': 'בוגרים'
                })
    return pd.DataFrame(second_table_rows)

# Function to update Goals table with User-Player IDs
def update_goals_table(file_path):
    df_goal_1 = pd.read_excel(file_path, sheet_name='Goals')
    df_user_player = pd.read_excel(file_path, sheet_name='Users-Players')
    merged_df = pd.merge(df_goal_1, df_user_player[['User_ID', 'Full_Name']], on='Full_Name', how='left')
    df_goal_1['User_ID'] = merged_df['User_ID_y']
    save_to_excel(df_goal_1, file_path, 'Goals')

# Function to update Referees table with matching User IDs
def update_referees_table(file_path):
    df_referees = pd.read_excel(file_path, sheet_name='Users-Referees')
    df_players = pd.read_excel(file_path, sheet_name='Users-Players')
    merged_df = pd.merge(df_referees, df_players[['User_ID', 'Full_Name']], on='Full_Name', how='left', suffixes=('_referee', '_player'))
    df_referees.loc[merged_df['User_ID_player'].notnull(), 'User_ID'] = merged_df['User_ID_player']
    save_to_excel(df_referees, file_path, 'Users-Referees')

# Function to update Games table with Referee IDs
def update_games_table(file_path):
    df_games = pd.read_excel(file_path, sheet_name='Games')
    df_users = pd.read_excel(file_path, sheet_name='Users')
    merged_df = pd.merge(df_games, df_users[['User_ID', 'Full_Name']], left_on='Referee', right_on='Full_Name', how='left')
    df_games['Referee_ID'] = merged_df['User_ID']
    save_to_excel(df_games, file_path, 'Games')

# Function to save DataFrame to Excel
def save_to_excel(df, file_path, sheet_name):
    book = load_workbook(file_path)
    if sheet_name in book.sheetnames:
        std = book[sheet_name]
        book.remove(std)
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        writer.book = book
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    print(f"Updated {sheet_name} table successfully saved to {file_path}")

# Function to convert Excel sheets to JSON
def excel_to_json(file_path, output_file):
    final = {}
    for sheet_name in pd.ExcelFile(file_path).sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        for col in df.select_dtypes(include=[np.number]).columns:
            df[col] = df[col].apply(lambda x: int(x) if pd.notnull(x) and (x == int(x)) else x)
        new = {}
        for row in range(len(df)):
            row_dict = {}
            for header in df.columns:
                value = df.at[row, header]
                if pd.isna(value):
                    row_dict[header] = None
                else:
                    if isinstance(value, (pd.Timestamp, datetime.datetime)):
                        row_dict[header] = value.isoformat()
                    elif isinstance(value, (pd.Timedelta, pd.Period)):
                        row_dict[header] = str(value)
                    elif isinstance(value, (np.integer, np.int64)):
                        row_dict[header] = int(value)
                    elif isinstance(value, (np.floating, np.float64)):
                        row_dict[header] = int(value) if value.is_integer() else float(value)
                    else:
                        row_dict[header] = value
            new[row] = row_dict
        final[sheet_name] = new
    with open(output_file, 'w', encoding='utf-8') as json_file:
        json.dump(final, json_file, cls=NpEncoder, ensure_ascii=False, indent=4)
    print("JSON file created successfully.")

# Main execution
def main():
    excel_file = 'to move.xlsx'
    first_table = pd.read_excel(excel_file, sheet_name='Goals to move').fillna(0)
    second_table = create_second_table(first_table)
    print(second_table)
    second_table.to_excel('Hocky.xlsx', sheet_name='Goals_1', index=False)

    file_path = 'Hockey.xlsx'
    update_goals_table(file_path)
    update_referees_table(file_path)
    update_games_table(file_path)
    excel_to_json(file_path, 'output.json')

if __name__ == "__main__":
    main()
